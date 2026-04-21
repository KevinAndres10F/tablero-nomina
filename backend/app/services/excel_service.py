"""Data service that reads payroll data from the bundled NOMINA_EJEMPLO.xlsx.

It exposes the same functions / response shape as the BigQuery and Google
Sheets services so `_get_data_service()` in main.py can swap it in.
"""

import os
import re
import threading
import time
from datetime import date, datetime
from functools import lru_cache
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook


FILTER_OPTIONS_TTL_SECONDS = int(os.getenv("FILTER_OPTIONS_TTL_SECONDS", "300"))
OVERVIEW_CACHE_TTL_SECONDS = int(os.getenv("OVERVIEW_CACHE_TTL_SECONDS", "45"))

_filter_options_cache: Optional[Dict[str, List[Any]]] = None
_filter_options_cached_at: float = 0.0
_filter_options_lock = threading.Lock()

_overview_cache: Dict[Tuple[Optional[str], Optional[int]], Tuple[float, Dict[str, Any]]] = {}
_overview_cache_lock = threading.Lock()


# Header row in NOMINA_EJEMPLO.xlsx is at row 3, data starts at row 4.
# Column B (index 1) holds NOMBRES; column A is empty.
_HEADER_ROW = 3
_DATA_START_ROW = 4

# Mapping from Excel header (column index inside the row tuple returned by
# openpyxl) to the canonical key the frontend expects.
_COLUMN_MAP: Dict[int, str] = {
    1: "nombre",
    2: "cedula",
    3: "genero",
    4: "fecha_nacimiento",
    5: "edad",
    6: "discapacidad",
    7: "porc_discapacidad",
    8: "tipo_contrato",
    9: "fecha_ingreso",
    10: "fecha_salida",
    11: "motivo_salida",
    12: "cargo",
    13: "centro_costo",
    14: "location",
    15: "class_",
    16: "area",            # CANAL DE VENTA se mapea como area.
    17: "mall",
    18: "provincia",
    19: "ciudad",
    20: "region",
    21: "jornada",
    22: "periodo",
    23: "dias_trabajados",
    24: "remun_unif",
    25: "horas_recnocturno",
    26: "recnocturno",
    27: "num_h_ext_100",
    28: "h_ext_100",
    29: "bono",
    30: "comisiones",
    31: "d14_mens_cos",
    32: "d14_mensual",
    33: "d13_mensual",
    34: "fondo_reserva_mensual",
    35: "total_ingresos",
    36: "pr_h_iess",
    37: "pr_q_iess",
    38: "iess_personal",
    39: "imp_renta",
    40: "seg_medico",
    41: "total_descuentos",
    42: "a_recibir",
    43: "ad_441_jp",
    44: "decimo_13",
    45: "decimo_14_c",
    46: "decimo_14",
    47: "fondos_reserva",
    48: "iess_patronal",
    49: "prov_bono",
    50: "vacaciones_prov",
    51: "total_provisiones",
}

_STRING_FIELDS = {
    "nombre", "cedula", "genero", "discapacidad", "tipo_contrato",
    "motivo_salida", "cargo", "centro_costo", "location", "class_",
    "area", "mall", "provincia", "ciudad", "region", "jornada",
}
_DATE_FIELDS = {"fecha_nacimiento", "fecha_ingreso", "fecha_salida"}


def _month_name(month: int) -> str:
    months = {
        1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
        5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
        9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
    }
    return months.get(month, "")


def _to_float(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    raw = str(value).strip().replace("$", "").replace(" ", "")
    if not raw:
        return 0.0
    if "," in raw and "." in raw:
        if raw.rfind(",") > raw.rfind("."):
            raw = raw.replace(".", "").replace(",", ".")
        else:
            raw = raw.replace(",", "")
    elif "," in raw:
        raw = raw.replace(",", ".")
    try:
        return float(raw)
    except ValueError:
        return 0.0


def _to_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _to_date_str(value: Any) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def _normalize_period(value: Any) -> str:
    raw = _to_text(value)
    if not raw:
        return ""
    match = re.match(r"^(\d{4})[-/](\d{1,2})$", raw)
    if match:
        return f"{int(match.group(1)):04d}-{int(match.group(2)):02d}"
    compact = re.match(r"^(\d{4})(\d{2})$", raw)
    if compact:
        return f"{int(compact.group(1)):04d}-{int(compact.group(2)):02d}"
    return raw


def _extract_year(periodo: str) -> Optional[int]:
    match = re.match(r"^(\d{4})", periodo or "")
    if not match:
        return None
    return int(match.group(1))


def _find_excel_path() -> Path:
    override = os.getenv("EXCEL_DATA_PATH", "").strip()
    if override:
        return Path(override).expanduser().resolve()

    here = Path(__file__).resolve()
    candidates = [
        here.parents[1] / "data" / "NOMINA_EJEMPLO.xlsx",   # backend/app/data/
        here.parents[2] / "NOMINA_EJEMPLO.xlsx",             # backend/
        here.parents[3] / "NOMINA_EJEMPLO.xlsx",             # repo root
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return candidates[0]


def _empty_row() -> Dict[str, Any]:
    row: Dict[str, Any] = {name: "" if name in _STRING_FIELDS or name in _DATE_FIELDS else 0.0
                          for name in _COLUMN_MAP.values()}
    row["periodo"] = ""
    row["edad"] = 0.0
    row["porc_discapacidad"] = 0.0
    return row


def _map_row(values: Tuple[Any, ...]) -> Dict[str, Any]:
    row = _empty_row()
    for idx, key in _COLUMN_MAP.items():
        if idx >= len(values):
            continue
        raw = values[idx]
        if key == "periodo":
            row[key] = _normalize_period(raw)
        elif key in _DATE_FIELDS:
            row[key] = _to_date_str(raw)
        elif key in _STRING_FIELDS:
            row[key] = _to_text(raw)
        else:
            row[key] = _to_float(raw)

    # Derivados que el frontend espera.
    row["h_ext_50"] = 0.0  # El archivo solo trae horas extra al 100%.
    row["horas_extras"] = row.get("h_ext_100", 0.0) + row.get("h_ext_50", 0.0)
    return row


def _has_data(row: Dict[str, Any]) -> bool:
    if not row.get("cedula") and not row.get("nombre"):
        return False
    return True


@lru_cache(maxsize=1)
def _load_rows_cached() -> List[Dict[str, Any]]:
    path = _find_excel_path()
    if not path.exists():
        raise FileNotFoundError(f"No se encontro el archivo Excel en {path}")

    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb.active
        rows: List[Dict[str, Any]] = []
        for idx, raw_row in enumerate(ws.iter_rows(values_only=True), start=1):
            if idx < _DATA_START_ROW:
                continue
            mapped = _map_row(raw_row)
            if _has_data(mapped):
                rows.append(mapped)
        return rows
    finally:
        wb.close()


def _get_rows() -> List[Dict[str, Any]]:
    return _load_rows_cached()


def _filter_rows(rows: List[Dict[str, Any]], periodo: Optional[str] = None,
                 year: Optional[int] = None) -> List[Dict[str, Any]]:
    if periodo:
        target = _normalize_period(periodo)
        return [r for r in rows if r.get("periodo") == target]
    if year:
        return [r for r in rows if _extract_year(r.get("periodo", "")) == int(year)]
    return rows


def _latest_row_by_employee(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    latest: Dict[str, Dict[str, Any]] = {}
    for row in rows:
        key = row.get("cedula") or row.get("nombre")
        if not key:
            continue
        current = latest.get(key)
        if not current or row.get("periodo", "") > current.get("periodo", ""):
            latest[key] = row
    return list(latest.values())


def fetch_filter_options() -> Dict[str, List[Any]]:
    global _filter_options_cache, _filter_options_cached_at
    with _filter_options_lock:
        now = time.time()
        if _filter_options_cache and (now - _filter_options_cached_at) < FILTER_OPTIONS_TTL_SECONDS:
            return _filter_options_cache

    rows = _get_rows()
    areas = sorted({r.get("area", "") for r in rows if r.get("area")})
    contratos = sorted({r.get("tipo_contrato", "") for r in rows if r.get("tipo_contrato")})

    periodos_unicos = sorted({r.get("periodo", "") for r in rows if r.get("periodo")}, reverse=True)
    periodos: List[Dict[str, Any]] = []
    for periodo in periodos_unicos:
        year = _extract_year(periodo)
        month = "01"
        label = periodo
        match = re.match(r"^(\d{4})-(\d{2})$", periodo)
        if match:
            month = match.group(2)
            label = f"{_month_name(int(month))} {match.group(1)}"
        periodos.append({
            "value": periodo,
            "year": str(year) if year else "",
            "month": month,
            "label": label,
        })

    years = sorted({y for y in (_extract_year(p) for p in periodos_unicos) if y is not None}, reverse=True)

    response = {
        "areas": areas,
        "contratos": contratos,
        "periodos": periodos,
        "years": years,
    }

    with _filter_options_lock:
        _filter_options_cache = response
        _filter_options_cached_at = time.time()

    return response


def fetch_employees(limit: int = 50, offset: int = 0, periodo: Optional[str] = None,
                    year: Optional[int] = None) -> Tuple[List[Dict[str, Any]], int]:
    rows = _filter_rows(_get_rows(), periodo=periodo, year=year)
    ordered = sorted(rows, key=lambda r: (r.get("periodo", ""), r.get("total_ingresos", 0.0)), reverse=True)
    total = len(ordered)
    sliced = ordered[offset: offset + limit]

    employees: List[Dict[str, Any]] = []
    for row in sliced:
        employee = dict(row)
        employee["horas_extras"] = employee.get("h_ext_100", 0.0) + employee.get("h_ext_50", 0.0)
        employees.append(employee)

    return employees, total


def _fetch_kpis(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    costo_total = sum(r.get("total_ingresos", 0.0) + r.get("total_provisiones", 0.0) for r in rows)
    total_provisiones = sum(r.get("total_provisiones", 0.0) for r in rows)
    total_horas_extras = sum(r.get("h_ext_100", 0.0) + r.get("h_ext_50", 0.0) for r in rows)
    total_neto = sum(r.get("a_recibir", 0.0) for r in rows)

    unique_ids = {r.get("cedula") for r in rows if r.get("cedula")}
    total_empleados = len(unique_ids) if unique_ids else len(rows)
    total_empleados = max(total_empleados, 1)

    return [
        {
            "label": "Costo Total Nomina",
            "value": costo_total,
            "subtitle": f"{total_empleados} empleados",
            "change_pct": 0.0,
            "trend": "up",
            "icon": "payments",
        },
        {
            "label": "Costo por Empleado",
            "value": costo_total / total_empleados,
            "subtitle": "Ingresos + Provisiones",
            "change_pct": 0.0,
            "trend": "up",
            "icon": "person",
        },
        {
            "label": "Total Provisiones",
            "value": total_provisiones,
            "subtitle": "D13 + D14 + Vac + F.Res",
            "change_pct": 0.0,
            "trend": "up",
            "icon": "savings",
        },
        {
            "label": "Horas Extras",
            "value": total_horas_extras,
            "subtitle": "100% + 50%",
            "change_pct": 0.0,
            "trend": "up",
            "icon": "schedule",
        },
        {
            "label": "Neto a Pagar",
            "value": total_neto,
            "subtitle": "Liquido empleados",
            "change_pct": 0.0,
            "trend": "up",
            "icon": "account_balance_wallet",
        },
    ]


def _fetch_cost_breakdown(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    decimo_13 = sum(r.get("decimo_13", 0.0) for r in rows)
    decimo_14 = sum(r.get("decimo_14", 0.0) for r in rows)
    vacaciones = sum(r.get("vacaciones_prov", 0.0) for r in rows)
    fondos_reserva = sum(r.get("fondos_reserva", 0.0) for r in rows)
    iess_patronal = sum(r.get("iess_patronal", 0.0) for r in rows)
    total = sum(r.get("total_provisiones", 0.0) for r in rows)

    divisor = total if total > 0 else 1.0
    return [
        {"label": "Decimo Tercero", "percent": round((decimo_13 / divisor) * 100, 1), "value": decimo_13},
        {"label": "Decimo Cuarto", "percent": round((decimo_14 / divisor) * 100, 1), "value": decimo_14},
        {"label": "Vacaciones", "percent": round((vacaciones / divisor) * 100, 1), "value": vacaciones},
        {"label": "Fondos Reserva", "percent": round((fondos_reserva / divisor) * 100, 1), "value": fondos_reserva},
        {"label": "IESS Patronal", "percent": round((iess_patronal / divisor) * 100, 1), "value": iess_patronal},
    ]


def _fetch_monthly_costs(rows: List[Dict[str, Any]], year: Optional[int] = None) -> List[Dict[str, Any]]:
    filtered = _filter_rows(rows, year=year)
    grouped: Dict[str, Dict[str, Any]] = {}

    for row in filtered:
        periodo = row.get("periodo") or "Sin periodo"
        bucket = grouped.setdefault(
            periodo,
            {
                "periodo": periodo,
                "month": periodo,
                "costo_total": 0.0,
                "ingresos": 0.0,
                "provisiones": 0.0,
                "descuentos": 0.0,
                "empleados": set(),
            },
        )
        bucket["ingresos"] += row.get("total_ingresos", 0.0)
        bucket["provisiones"] += row.get("total_provisiones", 0.0)
        bucket["descuentos"] += row.get("total_descuentos", 0.0)
        bucket["costo_total"] += row.get("total_ingresos", 0.0) + row.get("total_provisiones", 0.0)
        cedula = row.get("cedula")
        if cedula:
            bucket["empleados"].add(cedula)

    items: List[Dict[str, Any]] = []
    for periodo in sorted(grouped.keys()):
        bucket = grouped[periodo]
        match = re.match(r"^(\d{4})-(\d{2})$", periodo)
        if match:
            bucket["month"] = _month_name(int(match.group(2)))[:3]
        bucket["empleados"] = len(bucket["empleados"])
        items.append(bucket)

    return items


def _fetch_distribution_by_contract(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    latest_rows = _latest_row_by_employee(rows)
    grouped: Dict[str, Dict[str, Any]] = {}
    for row in latest_rows:
        label = row.get("tipo_contrato") or "Sin definir"
        bucket = grouped.setdefault(label, {"label": label, "count": 0, "value": 0.0})
        bucket["count"] += 1
        bucket["value"] += row.get("total_ingresos", 0.0)
    return sorted(grouped.values(), key=lambda item: item["count"], reverse=True)


def _fetch_distribution_by_area(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    latest_rows = _latest_row_by_employee(rows)
    grouped: Dict[str, Dict[str, Any]] = {}
    for row in latest_rows:
        label = row.get("area") or "Sin definir"
        bucket = grouped.setdefault(label, {"label": label, "count": 0, "value": 0.0})
        bucket["count"] += 1
        bucket["value"] += row.get("total_ingresos", 0.0)
    return sorted(grouped.values(), key=lambda item: item["count"], reverse=True)


def fetch_overview(periodo: Optional[str] = None, year: Optional[int] = None) -> Dict[str, Any]:
    cache_key = (periodo, year)
    with _overview_cache_lock:
        cached_item = _overview_cache.get(cache_key)
        if cached_item and (time.time() - cached_item[0]) < OVERVIEW_CACHE_TTL_SECONDS:
            return cached_item[1]

    try:
        rows = _get_rows()
        filter_options = fetch_filter_options()

        effective_year = year
        if not effective_year and not periodo:
            years = filter_options.get("years", [])
            if years:
                effective_year = years[0]

        active_rows = _filter_rows(rows, periodo=periodo, year=effective_year)
        kpis = _fetch_kpis(active_rows)
        breakdown = _fetch_cost_breakdown(active_rows)
        monthly = _fetch_monthly_costs(rows, year=effective_year)
        employees, total_employees = fetch_employees(limit=500, periodo=periodo, year=effective_year)
        dist_contrato = _fetch_distribution_by_contract(active_rows)
        dist_area = _fetch_distribution_by_area(active_rows)

        if periodo:
            display_period = _normalize_period(periodo)
        elif effective_year:
            display_period = str(effective_year)
        else:
            display_period = "Todos los anos"

        response = {
            "kpis": kpis,
            "monthly_costs": monthly,
            "cost_breakdown": breakdown,
            "employees": employees,
            "distribution_contrato": dist_contrato,
            "distribution_area": dist_area,
            "currency": "USD",
            "period": display_period,
            "total_employees": total_employees,
            "filter_options": filter_options,
            "selected_year": effective_year,
            "selected_periodo": _normalize_period(periodo) if periodo else None,
        }

        with _overview_cache_lock:
            _overview_cache[cache_key] = (time.time(), response)

        return response
    except Exception as exc:
        return {
            "kpis": [],
            "monthly_costs": [],
            "cost_breakdown": [],
            "employees": [],
            "distribution_contrato": [],
            "distribution_area": [],
            "currency": "USD",
            "period": "N/D",
            "total_employees": 0,
            "filter_options": {"areas": [], "contratos": [], "periodos": [], "years": []},
            "error": str(exc),
        }
